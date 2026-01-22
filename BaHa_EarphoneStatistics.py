import requests
from bs4 import BeautifulSoup
import re
import requests
import pandas as pd
import os
import random
import logging
from selenium import webdriver 
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

URL = "https://forum.gamer.com.tw/C.php?page=1&bsn=60535&snA=28366"
User_Agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:119.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
]
# 初始 headers 使用隨機 UA
headers = {"User-Agent": None}
soup = None
post_containers = []
total_pages = 0
rows = []
today_str = pd.Timestamp.now().strftime('%Y%m%d')
# 初始化 Selenium WebDriver (以 Edge 為例)
opt = webdriver.EdgeOptions()
opt.add_argument('--headless=new')
# 關閉日誌輸出
opt.add_argument('--log-level=3')
opt.add_argument('--disable-logging')
driver = webdriver.Edge(options=opt)

# logging
logger = logging.getLogger(__name__)
log_level = logging.DEBUG if os.getenv('BAHA_DEBUG', '').lower() in ('1', 'true', 'yes') else logging.INFO
logging.basicConfig(level=log_level, format='%(asctime)s %(levelname)s: %(message)s')

# 初始化頁面
def init_page(pageURL):
    headers["User-Agent"] = get_random_ua(exclude=headers.get("User-Agent"))
    request_status = requests.get(pageURL, headers=headers)

    if request_status.status_code == 200:
        set_page(request_status_text=request_status.text)
        try:
            # 在 Selenium 中也載入該頁面，之後可用來執行 hover 操作
            driver.get(pageURL)
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'article.c-post'))
            )
            time.sleep(0.3)
        except Exception:
            pass

    else:
        # 請求次數
        try_times = 5
        # 當前嘗試次數
        current_try = 0
        
        while request_status.status_code != 200 and current_try < try_times:
            headers["User-Agent"] = get_random_ua(exclude=headers.get("User-Agent"))
            request_status = requests.get(URL, headers=headers)

            if request_status.status_code == 200:
                set_page(request_status_text=request_status.text)
                break
            else:
                current_try += 1
                print(f"錯誤狀態碼: {request_status.status_code}")
                print(f"重新要求... {current_try}/{try_times}")

        if current_try == try_times and request_status.status_code != 200:
            print("錯誤要求多次，請稍後再試...")
            exit(1)

# 嘗試多次載入頁面，先用 requests，失敗後用 Selenium
def fetch_page_with_retry(page_url, max_attempts=3, selenium_fallback=True):

    global post_containers
    for attempt in range(1, max_attempts + 1):
        try:
            logger.debug('Fetch attempt %d for %s via requests', attempt, page_url)
            init_page(page_url)
        except Exception as e:
            logger.exception('init_page failed on attempt %d: %s', attempt, e)

        if post_containers and len(post_containers) > 0:
            logger.debug('Found %d posts via requests on attempt %d', len(post_containers), attempt)
            return True

        # requests didn't yield posts — try Selenium render if allowed
        if selenium_fallback:
            try:
                logger.info('Requests returned no posts; trying Selenium render for %s (attempt %d)', page_url, attempt)
                driver.get(page_url)
                time.sleep(0.6)
                set_page(driver.page_source)
            except Exception as e:
                logger.exception('Selenium fetch failed on attempt %d: %s', attempt, e)

            if post_containers and len(post_containers) > 0:
                logger.debug('Found %d posts via Selenium on attempt %d', len(post_containers), attempt)
                return True

        # small backoff before next attempt
        time.sleep(0.5 * attempt)

    logger.warning('No posts found for %s after %d attempts', page_url, max_attempts)
    return False

# 設定頁面
def set_page(request_status_text):
    global soup, post_containers
    soup = BeautifulSoup(request_status_text, 'html.parser')
    post_containers = soup.select('.c-post')   

# 回傳隨機 UA，排除指定的 UA(exclude)
def get_random_ua(exclude=None):
    if exclude is None:
        return random.choice(User_Agents)
    if len(User_Agents) == 1:
        return User_Agents[0]
    ua = random.choice(User_Agents)
    tries = 0
    while ua == exclude and tries < 10:
        ua = random.choice(User_Agents)
        tries += 1
    if ua == exclude:
        for u in User_Agents:
            if u != exclude:
                return u
    return ua

# 解析內容區塊
def parse_content_sections(content):

    if not content:
        return {'耳罩': '', '耳塞': '', '前端': ''}

    sections = {'耳罩': [], '耳塞': [], '前端': []}
    text = content.strip()

    # 支援多種別名，將別名對回 canonical 欄位
    alias_map = {
        '耳罩': ['耳罩', '耳罩式耳機', '耳罩式', '大耳'],
        '耳塞': ['耳塞', '耳道', '入耳', '耳道式', '耳道式耳機', '公模', '私模', '塞子'],
        '前端': ['前端', 'DAP', '小尾巴', '隨身前端', 'DAC/AMP/一體機/DDC', 'CD機', '大件', '設備', '播放器', '系統', '一體機', '前端(訊號處理)']
    }
    # 建立別名到 canonical 的反向索引
    alias_to_key = {}
    for k, aliases in alias_map.items():
        for a in aliases:
            alias_to_key[a] = k

    # 為了避免較短別名搶先匹配，依長度排序（長的先匹配）
    sorted_aliases = sorted(alias_to_key.keys(), key=lambda x: -len(x))

    # 先用正規式抓取標題 + 內容直到下一個標題（支援所有別名）
    alias_group = '|'.join(re.escape(a) for a in sorted_aliases)
    pattern = re.compile(rf'({alias_group})\s*[:：\-–]?\s*(.*?)(?=({alias_group})|\Z)', re.S)
    matches = pattern.findall(text)
    if matches:
        for m in matches:
            raw_key = m[0]
            key = alias_to_key.get(raw_key, raw_key)
            val = m[1].strip()
            if val:
                val = '\n'.join([ln.strip() for ln in val.splitlines() if ln.strip()])
                sections[key].append(val)

        return {k: ','.join(v) if v else '' for k, v in sections.items()}         
    
    return {'耳罩': '', '耳塞': '', '前端': ''}

# 最後整理：保留原始合併欄位，同時為每個 section 產生 廠牌 與 型號 欄位
def split_brand_model(s: str):
    s = s.strip()

    if not s:
        return ('', '')
    
    # 常見分隔符號（注意：不以純空白作為主要分隔符，避免把型號切碎）
    seps = [',', '，', '/', '／', ' - ', ' – ', ' -', '-']
    for sep in seps:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) >= 2:
                brand = parts[0].upper()
                model = parts[1]
                return (brand, model)
    # 若找不到分隔符，預設視為廠牌（並轉大寫）
    return (s.upper(), '')

# 取得總頁數
def get_total_pages():
    # 嘗試用 requests 多次取得並解析分頁，若失敗則使用 Selenium 做 fallback
    if headers["User-Agent"] is None:
        headers["User-Agent"] = get_random_ua()

    else:
        headers["User-Agent"] = get_random_ua(exclude=headers.get("User-Agent"))


    for attempt in range(3):
        try:
            r = requests.get(URL, headers=headers, timeout=8)
        except Exception:
            time.sleep(0.5)
            continue

        if r.status_code != 200:
            time.sleep(0.5)
            continue

        soup_local = BeautifulSoup(r.text, 'html.parser')

        # 先找常見的分頁容器
        pagination = soup_local.find('p', class_='BH-pagebtnA') or soup_local.find('div', class_='BH-pagebtnA')
        page_numbers = []

        if pagination:
            for a in pagination.find_all('a'):
                txt = a.get_text().strip()
                if txt.isdigit():
                    page_numbers.append(int(txt))
                else:
                    href = a.get('href', '')
                    m = re.search(r'[?&]page=(\d+)', href)
                    if m:
                        page_numbers.append(int(m.group(1)))

        # 若仍找不到，可掃描頁面所有連結尋找 page= 的 href 或數字文字
        if not page_numbers:
            for a in soup_local.find_all('a', href=True):
                href = a['href']
                m = re.search(r'[?&]page=(\d+)', href)
                if m:
                    page_numbers.append(int(m.group(1)))
                else:
                    txt = a.get_text().strip()
                    if txt.isdigit():
                        page_numbers.append(int(txt))

        if page_numbers:
            total = max(page_numbers)
            if total >= 1:
                return total

    # requests 無法正確取得多頁資訊，使用 Selenium 做最後嘗試（rendered page）
    try:
        headers["User-Agent"] = get_random_ua(exclude=headers.get("User-Agent"))
        driver.get(URL)
        time.sleep(0.5)
        els = driver.find_elements(By.CSS_SELECTOR, 'p.BH-pagebtnA a, div.BH-pagebtnA a, a')
        page_numbers = []
        for el in els:
            txt = el.text.strip()
            try:
                if txt.isdigit():
                    page_numbers.append(int(txt))
            except Exception:
                pass
            try:
                href = el.get_attribute('href') or ''
                m = re.search(r'[?&]page=(\d+)', href)
                if m:
                    page_numbers.append(int(m.group(1)))
            except Exception:
                pass

        if page_numbers:
            return max(page_numbers)
    except Exception:
        pass

    # 最後退回 1（呼叫端須處理只有 1 的情境）
    return 1

# Excel初始化格式
def excel_init(out, df):
    # 寫入 Excel，並將 Link 欄設為超連結（優先使用 openpyxl）
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        import unicodedata

        sheet_name = '【原始資料】2026年第六屆巴哈影音普查'
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        #region 第一列：合併欄位

        ws.merge_cells('A1:A2')
        excel_title_style(ws['A1'])
        ws['A1'] = '名稱(樓層)'
        ws.merge_cells('B1:D1')
        excel_title_style(ws['B1'])
        ws['B1'] = '內容'
        ws.merge_cells('E1:E2')
        excel_title_style(ws['E1'])
        ws['E1'] = '連結'

        #endregion  

        #region 第二列：子欄位（每個區塊拆成 廠牌, 型號）

        subtitle_row = [ '耳罩', '耳塞', '前端']

        for col_idx, h in enumerate(subtitle_row, start=2):
            ws.cell(row=2, column=col_idx, value=h)
            excel_title_style(ws.cell(row=2, column=col_idx))

        #region 寫入資料，從第3列開始

        for i, row in enumerate(df.to_dict(orient='records'), start=3):
            # 名稱(樓層)
            name_floor = f"{row.get('名稱', '')}({row.get('樓層', '')})"
            ws.cell(row=i, column=1, value=name_floor)
            # 每個區塊拆成 廠牌, 型號
            ws.cell(row=i, column=2, value=row.get('耳罩', ''))
            ws.cell(row=i, column=3, value=row.get('耳塞', ''))
            ws.cell(row=i, column=4, value=row.get('前端', ''))

            link_val = row.get('連結', '')
            if link_val:
                cell = ws.cell(row=i, column=5, value='連結')
                cell.hyperlink = link_val
                cell.style = 'Hyperlink'
            else:
                ws.cell(row=i, column=5, value='')
        #endregion

        #region 自動調整欄寬

        def visual_width(s: str) -> int:

            w = 0
            for ch in s:
                # east_asian_width 回傳 'F','W','A','N','Na','H'，視為寬字的通常是 F or W
                ea = unicodedata.east_asian_width(ch)
                if ea in ('F', 'W'):
                    w += 2
                else:
                    w += 1
            return w

        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_w = 0
            for cell in ws[col_letter]:
                if cell.value is None:
                    continue
                # 考慮換行：取換行後每行的最大 visual_width
                lines = str(cell.value).splitlines()
                longest = 0
                for ln in lines:
                    lw = visual_width(ln)
                    if lw > longest:
                        longest = lw
                if longest > max_w:
                    max_w = longest
            # 將字元數轉為 Excel 欄寬近似值，加上 padding
            # 這裡不做更複雜字型換算，直接使用字元寬度近似
            if max_w <= 0:
                width = 8
            else:
                width = min(int(max_w) + 2, 120)
            ws.column_dimensions[col_letter].width = width

        # 根據內容行數調整列高（第3列開始）
        for row_idx in range(3, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(2, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    lines = str(cell.value).splitlines()
                    if len(lines) > max_lines:
                        max_lines = len(lines)
            ws.row_dimensions[row_idx].height = max_lines * 15

        #endregion

        wb.save(out)

        print(f"已輸出至 {out}")

    except Exception:
        # fallback: 寫入 HYPERLINK 公式（若 openpyxl 不可用）
        df2 = df.copy()

        df2.to_excel(out, index=False)

        print(f"已輸出至 {out}（不分樣式）")

# Excel 標題樣式
def excel_title_style(ws):
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    thin = Side(border_style="thin", color="000000")

    ws.alignment = Alignment(horizontal='center', vertical='center')
    ws.fill = PatternFill(start_color='7BBBFE', end_color='7BBBFE', fill_type='solid')
    ws.font = Font(bold=True)
    ws.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 複製 手動調整 sheet
def copy_sheet_manual(out_path, new_name="【手動調整】2026年第六屆巴哈影音普查"):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(out_path)
        # 使用第一個工作表作為來源
        src = wb[wb.sheetnames[0]]
        copied = wb.copy_worksheet(src)

        # 若名稱衝突，加入遞增後綴
        base = new_name
        i = 1
        while new_name in wb.sheetnames:
            new_name = f"{base}_{i}"
            i += 1
        copied.title = new_name
        wb.save(out_path)

    except Exception as e:
        print("複製工作表失敗:", e)

total_pages = get_total_pages()

for page in range(1, total_pages + 1):
    print(f"正在處理第 {page} 頁，共 {total_pages} 頁")
    page_url = f"https://forum.gamer.com.tw/C.php?page={page}&bsn=60535&snA=28366"
    headers["User-Agent"] = get_random_ua(exclude=headers.get("User-Agent"))
    ok = fetch_page_with_retry(page_url, max_attempts=3, selenium_fallback=True)
    if not ok:
        logger.warning('Skipping page %d because no posts were found after retries', page)
        continue

    selenium_posts = driver.find_elements(By.CLASS_NAME, 'c-post')

    for idx, pc in enumerate(post_containers):
        username = ''
        floor = ''
        content = ''
        link = ''

        #region 取得使用者名稱

        username_el = pc.select_one('a.username')
        if username_el:
            username = username_el.get_text(strip=True)
        
        #endregion

        #region 取得樓層

        floor_el = pc.select_one('a.floor')
        if floor_el:
            floor = floor_el.get_text(strip=True)

        if floor == '樓主':
            continue  # 跳過樓主  

        #endregion

        #region 取得內容，保留換行
        content_el = pc.select_one('.c-article__content') or pc.select_one('.c-post__body') or pc.select_one('.postContent') or pc

        for bad in content_el.select('.c-article__meta, .c-article__author, .c-article__info, .meta, .tools, a.copy-link'):
            bad.decompose()

        content = content_el.get_text('\n', strip=True)

        #endregion

        #region 取得連結

        try:
            sp = None
            if idx < len(selenium_posts):
                sp = selenium_posts[idx]
            elif selenium_posts:
                sp = selenium_posts[-1]

            if sp is not None:
                # 嘗試 hover 特定 menu 元素；若找不到則 hover 整個文章元素
                try:
                    menu = sp.find_element(By.CLASS_NAME, 'tippy-option-menu')
                    ActionChains(driver).move_to_element(menu).perform()
                except Exception:
                    try:
                        ActionChains(driver).move_to_element(sp).perform()
                    except Exception:
                        pass

                # 等待並取得 copy-link
                try:
                    btn = WebDriverWait(sp, 1).until(
                        lambda el: el.find_element(By.CSS_SELECTOR, 'a[data-action="copyLink"]')
                    )

                    driver.execute_script("""
                        window._copied_value = null;
                        try {
                        if (!navigator.clipboard) navigator.clipboard = {};
                        navigator.clipboard.writeText = function(text) { window._copied_value = text; return Promise.resolve(); };
                        } catch (e) { window._copied_value = null; }
                        """)
                    
                    # 觸發按鈕 click
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.2)
                    copied = driver.execute_script("return window._copied_value || null;")

                    if copied:
                        link = copied
                    
                except Exception:
                        link = ''
        except Exception:
            link = ''
            

        #endregion

        parts = parse_content_sections(content)
        rows.append({
            '名稱': username,
            '樓層': floor,
            '內容': content,
            '耳罩': parts.get('耳罩', ''),
            '耳塞': parts.get('耳塞', ''),
            '前端': parts.get('前端', ''),
            '連結': link
        })

#region 輸出成 Excel 檔案

# 創建資料夾
out_dir = os.path.join(os.path.dirname(__file__), 'output')
os.makedirs(out_dir, exist_ok=True)

# # 輸出成Excel
df = pd.DataFrame(rows)
out = os.path.join(out_dir, f'BaHa_EarphoneStatistics_{today_str}.xlsx')

excel_init(out, df)

copy_sheet_manual(out)

#endregion
